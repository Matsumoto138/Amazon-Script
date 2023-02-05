import os
# Import Excel Library
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

#Import Selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service

#Import zipfile
from zipfile import ZipFile

#Import Json
import json
import time

# Excel dosyası aç
wb = load_workbook('Amazon Order.xlsx')
wb2 = Workbook()
ws = wb.active
ws2 = wb2.active
ws2.title = "Data"

# Değişkenleri tanımla
col_data = []
col_order = []
new_file_col = []
col_url = ""
counter = 0
title = ""
color = ""
length = ""
customization = ""


# drive tanımla
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
prefs = {"profile.default_content_settings.popups": 0,
             "download.default_directory": 
                        os.path.dirname(__file__),
             "directory_upgrade": False}
options.add_experimental_option("prefs", prefs)
driver=webdriver.Chrome("./chromedriver/chromedriver.exe", options=options)


# Url al ve col data'ya ekle
for row in range(2, ws.max_row + 1):
    col_order.append(ws["B"+ str(row)].value)
    col_data.append(ws["AI" + str(row)].value)
    

# Dosya Loop
for i in col_data:

    #dosyaları indir
    try:
        driver.get(i)
    except:
        ws2.append(["","","",""])
        continue
    time.sleep(3)

    # .zip dosyasını bul ve içindekileri çıkart
    for file in os.listdir(os.path.dirname(__file__)):
        
        if file.endswith(".zip"):
            
            with ZipFile(file, 'r') as zip:
                zip.extractall()
            
    # Çıkartılan dosyalar içinde .json dosyasını bul ve bilgileri diziye ekle
    for fileJson in os.listdir(os.path.dirname(__file__)):
        if fileJson.endswith(".json"):
            
            name = os.path.basename(fileJson)
            jsonName = os.path.splitext(name)[0]
            
            f = open(jsonName+".json")
            data = json.load(f)

            try:
                title = data["title"]
                color= data["customizationData"]["children"][0]["children"][0]["optionSelection"]["label"]
                length= data["customizationData"]["children"][0]["children"][1]["optionSelection"]["label"]
                customization= data["customizationData"]["children"][0]["children"][2]["children"][0]["inputValue"]
            except:
                ws2.append([title,"","",""])
                f.close()
                continue

            ws2.append([title, color, length, customization])
            wb2.save("Amazon New Order.xlsx")
            f.close()
            time.sleep(3)
            
    for deleteFile in os.listdir(os.path.dirname(__file__)):
        if deleteFile.endswith(".json") or deleteFile.endswith(".zip") or deleteFile.endswith(".xml"):
            try:
                os.remove(deleteFile)
            except:
                time.sleep(3)
    
